# apps/campaigns/views.py

from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.utils import timezone
from django.core.mail import EmailMultiAlternatives
from django.http import HttpResponse, HttpResponseRedirect
import csv
from django.views.decorators.http import require_GET
from django.http import Http404
import uuid
from django.db import transaction

from .models import Campaign, CampaignStats, EmailTracking, CampaignRecipient
from .serializers import CampaignSerializer, CampaignListSerializer
from apps.billing.models import Plan
from apps.campaigns.tasks import send_campaign
from django.conf import settings


class CampaignViewSet(viewsets.ModelViewSet):
    """
    REST API: CRUD кампаний + экшен send.
    """
    serializer_class = CampaignSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.is_authenticated:
            return Campaign.objects.none()
        return Campaign.objects.filter(user=self.request.user).order_by('-created_at')

    def get_serializer_class(self):
        if self.action == 'list':
            return CampaignListSerializer
        return CampaignSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        # Пагинация через page/page_size
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        total = queryset.count()
        num_pages = (total + page_size - 1) // page_size
        start = (page - 1) * page_size
        end = start + page_size
        page_qs = queryset[start:end]
        serializer = self.get_serializer(page_qs, many=True)
        return Response({
            'results': serializer.data,
            'count': total,
            'page': page,
            'num_pages': num_pages,
        })

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def retrieve(self, request, *args, **kwargs):
        # Принудительно обновляем данные из базы, игнорируя кэш
        instance = self.get_object()
        # Очищаем кэш для конкретной кампании
        cache_key = f"campaign_{instance.id}"
        from django.core.cache import cache
        cache.delete(cache_key)
        
        return super().retrieve(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        print(f"Update request data: {request.data}")
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        if not serializer.is_valid():
            print(f"Serializer errors: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        print(f"Serializer validated data: {serializer.validated_data}")
        
        # Если это черновик, сохраняем все поля как есть
        if instance.status == Campaign.STATUS_DRAFT:
            self.perform_update(serializer)
            print(f"Campaign after update: template={instance.template}, sender_email={instance.sender_email}, contact_lists={list(instance.contact_lists.all())}")
            return Response(serializer.data)
            
        # Для других статусов выполняем полную валидацию
        self.perform_update(serializer)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        """
        Защита от удаления активных кампаний.
        """
        instance = self.get_object()
        
        # Запрещаем удаление кампаний в процессе отправки
        if instance.status in [Campaign.STATUS_SENDING, Campaign.STATUS_PENDING]:
            return Response({
                'detail': f'Нельзя удалить кампанию со статусом "{instance.get_status_display()}". '
                         f'Дождитесь завершения отправки или отмените кампанию.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Если есть активная Celery задача, проверяем её статус
        if instance.celery_task_id:
            from celery.result import AsyncResult
            task_result = AsyncResult(instance.celery_task_id)
            
            if task_result.state in ['PENDING', 'STARTED']:
                return Response({
                    'detail': 'Нельзя удалить кампанию, которая сейчас отправляется. '
                             f'Статус задачи: {task_result.state}'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='cancel')
    def cancel(self, request, pk=None):
        """
        Отмена кампании (только для отправляющихся кампаний).
        """
        campaign = self.get_object()
        
        if campaign.status not in [Campaign.STATUS_SENDING, Campaign.STATUS_PENDING]:
            return Response({
                'detail': 'Можно отменить только отправляющиеся кампании.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Отменяем Celery задачу если она есть
        if campaign.celery_task_id:
            from celery.result import AsyncResult
            task_result = AsyncResult(campaign.celery_task_id)
            task_result.revoke(terminate=True)
        
        # Обновляем статус кампании
        campaign.status = Campaign.STATUS_FAILED
        campaign.celery_task_id = None
        campaign.save(update_fields=['status', 'celery_task_id', 'updated_at'])
        
        return Response({
            'detail': 'Кампания отменена.',
            'status': campaign.status
        })

    @action(detail=True, methods=['post'], url_path='send')
    def send(self, request, pk=None):
        campaign = self.get_object()
        
        print(f"Send campaign: {campaign.id}")
        print(f"Campaign status: {campaign.status}")
        print(f"Campaign template: {campaign.template}")
        print(f"Campaign sender_email: {campaign.sender_email}")
        print(f"Campaign contact_lists: {list(campaign.contact_lists.all())}")
        print(f"Campaign sender_name: {campaign.sender_name}")

        if campaign.status != Campaign.STATUS_DRAFT:
            return Response({'detail': 'Можно отправить только черновик.'},
                            status=status.HTTP_400_BAD_REQUEST)

        # Проверяем, не запущена ли уже кампания
        if campaign.celery_task_id:
            # Проверяем статус существующей задачи
            from celery.result import AsyncResult
            task_result = AsyncResult(campaign.celery_task_id)
            
            if task_result.state in ['PENDING', 'STARTED']:
                return Response({
                    'detail': 'Кампания уже отправляется.',
                    'task_id': campaign.celery_task_id,
                    'task_status': task_result.state
                }, status=status.HTTP_400_BAD_REQUEST)
            elif task_result.state in ['SUCCESS', 'FAILURE', 'REVOKED']:
                # Очищаем старый task_id
                campaign.celery_task_id = None
                campaign.save(update_fields=['celery_task_id'])

        # --- ОГРАНИЧЕНИЕ ПО ТАРИФУ НА ОТПРАВКУ ---
        user = request.user
        
        # Используем правильные функции из billing.utils
        from apps.billing.utils import can_user_send_emails, get_user_plan_info
        from apps.mailer.models import Contact as MailerContact
        
        # Сколько писем будет отправлено в этой кампании (считаем только VALID контакты)
        recipients_count = 0
        for cl in campaign.contact_lists.all():
            recipients_count += cl.contacts.filter(status=MailerContact.VALID).count()
        
        print(f"Recipients count: {recipients_count}")
        
        if recipients_count == 0:
            return Response({
                'error': 'В выбранных списках нет валидных контактов для отправки.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Получаем информацию о тарифе пользователя
        plan_info = get_user_plan_info(user)
        print(f"Plan info: {plan_info}")
        
        # Проверяем лимиты в зависимости от типа тарифа
        if plan_info['has_plan'] and plan_info['plan_type'] == 'Letters':
            # Тариф с письмами - проверяем остаток писем
            if not can_user_send_emails(user, recipients_count):
                remaining = plan_info.get('emails_remaining', 0)
                return Response({
                    'error': f'Количество получателей ({recipients_count}) превышает остаток писем в тарифе ({remaining}). Пожалуйста, пополните баланс или уменьшите количество получателей.'
                }, status=status.HTTP_400_BAD_REQUEST)
        elif plan_info['has_plan'] and plan_info['plan_type'] == 'Subscribers':
            # Тариф с подписчиками - проверяем срок действия и лимит подписчиков
            if plan_info['is_expired']:
                return Response({
                    'error': 'Тариф истёк. Пожалуйста, продлите тариф для отправки кампаний.'
                }, status=status.HTTP_400_BAD_REQUEST)
            # Проверяем лимит по подписчикам (если есть)
            if plan_info.get('subscribers_limit') and recipients_count > plan_info['subscribers_limit']:
                return Response({
                    'error': f'Количество получателей ({recipients_count}) превышает лимит тарифа ({plan_info["subscribers_limit"]} подписчиков). Пожалуйста, обновите тариф или уменьшите количество получателей.'
                }, status=status.HTTP_400_BAD_REQUEST)
        else:
            # Бесплатный тариф - проверяем лимиты
            if not can_user_send_emails(user, recipients_count):
                from apps.billing.models import BillingSettings
                free_limit = BillingSettings.get_settings().free_plan_subscribers or 200
                remaining = max(0, free_limit - (plan_info.get('emails_sent', 0) or 0))
                return Response({
                    'error': f'Превышен лимит бесплатного тарифа. Отправлено: {plan_info.get("emails_sent", 0)}, Лимит: {free_limit}, Остаток: {remaining}. Для отправки {recipients_count} писем необходимо приобрести платный тариф.'
                }, status=status.HTTP_400_BAD_REQUEST)

        # --- ВАЛИДАЦИЯ ПОЛЕЙ КАМПАНИИ ---
        print("Validating campaign fields:")
        print(f"  name: {campaign.name}")
        print(f"  template: {campaign.template}")
        print(f"  sender_email: {campaign.sender_email}")
        print(f"  subject: {campaign.subject}")

        # Проверяем обязательные поля
        if not campaign.name:
            return Response({'error': 'Название кампании обязательно.'}, status=status.HTTP_400_BAD_REQUEST)
        if not campaign.template:
            return Response({'error': 'Шаблон письма обязателен.'}, status=status.HTTP_400_BAD_REQUEST)
        if not campaign.sender_email:
            return Response({'error': 'Email отправителя обязателен.'}, status=status.HTTP_400_BAD_REQUEST)
        if not campaign.subject:
            return Response({'error': 'Тема письма обязательна.'}, status=status.HTTP_400_BAD_REQUEST)
        if not campaign.contact_lists.exists():
            return Response({'error': 'Список контактов обязателен.'}, status=status.HTTP_400_BAD_REQUEST)

        # Проверяем, что email отправителя верифицирован
        if not campaign.sender_email.is_verified:
            return Response({'error': 'Email отправителя должен быть верифицирован.'}, status=status.HTTP_400_BAD_REQUEST)

        print("Starting campaign sending...")
        
        try:
            # Проверяем доступность Celery
            from celery import current_app
            inspect = current_app.control.inspect()
            stats = inspect.stats()
            
            if not stats:
                return Response({
                    'error': 'Celery worker недоступен. Проверьте, что Celery запущен.',
                    'detail': 'No active workers found'
                }, status=status.HTTP_503_SERVICE_UNAVAILABLE)
            
            print(f"Found {len(stats)} active Celery workers")
            
            # Запускаем задачу Celery с дополнительными параметрами
            print("Starting Celery task...")
            # Условия безопасного пропуска модерации:
            # - доверенный или персонал
            # - небольшая рассылка (до 50 получателей)
            skip_moderation = (
                getattr(request.user, 'is_trusted_user', False)
                or getattr(request.user, 'is_staff', False)
                or recipients_count <= 50
            )

            # Сразу переводим кампанию в статус "sending"
            # и откладываем постановку задачи Celery до коммита транзакции
            def _enqueue_after_commit():
                try:
                    print("Starting Celery task...")
                    task = send_campaign.apply_async(
                        args=[campaign.id],
                        kwargs={'skip_moderation': skip_moderation},
                        queue='campaigns',
                        countdown=1,
                        expires=1800,
                        retry=True,
                        retry_policy={
                            'max_retries': 3,
                            'interval_start': 0,
                            'interval_step': 0.2,
                            'interval_max': 0.2,
                        }
                    )
                    print(f"Celery task started: {task.id}")
                    # Сохраняем task_id уже после коммита
                    try:
                        Campaign.objects.filter(id=campaign.id).update(celery_task_id=task.id)
                    except Exception as e:
                        print(f"Warning: failed to persist celery_task_id after commit: {e}")
                except Exception as e:
                    print(f"Error scheduling Celery task after commit: {e}")

            transaction.on_commit(_enqueue_after_commit)
            
            # Обновляем кампанию немедленно: статус -> sending
            try:
                with transaction.atomic():
                    campaign.status = Campaign.STATUS_SENDING
                    campaign.save(update_fields=['status', 'updated_at'])
            except Exception as e:
                print(f"Warning: failed to update campaign status to sending: {e}")
            
            return Response({
                'message': 'Кампания запущена на отправку.',
                # task_id и task_status будут записаны после коммита транзакции
                'task_id': None,
                'task_status': 'QUEUED',
                'recipients_count': recipients_count,
                'workers_count': len(stats)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            print(f"Error starting campaign: {e}")
            import traceback
            traceback.print_exc()
            
            # Очищаем task_id в случае ошибки
            campaign.celery_task_id = None
            campaign.save(update_fields=['celery_task_id'])
            
            return Response({
                'error': 'Ошибка запуска кампании.',
                'detail': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def progress(self, request, pk=None):
        """Получить прогресс отправки кампании"""
        campaign = self.get_object()
        
        from django.core.cache import cache
        
        # Получаем прогресс из кэша
        progress_data = cache.get(f'campaign_progress_{campaign.id}')
        
        if not progress_data:
            # Если нет в кэше, считаем из базы
            total_recipients = CampaignRecipient.objects.filter(campaign=campaign).count()
            sent_recipients = CampaignRecipient.objects.filter(
                campaign=campaign, 
                is_sent=True
            ).count()
            
            progress_data = {
                'total': total_recipients,
                'sent': sent_recipients,
                'progress': (sent_recipients / total_recipients * 100) if total_recipients > 0 else 0
            }
        
        return Response({
            'campaign_id': str(campaign.id),
            'status': campaign.status,
            'progress': progress_data
        })

    @action(detail=True, methods=['post'])
    def track_open(self, request, pk=None):
        """Отслеживание открытия письма"""
        tracking_id = request.GET.get('tracking_id')
        if not tracking_id:
            return Response({'error': 'Tracking ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            tracking = EmailTracking.objects.get(
                campaign_id=pk,
                tracking_id=tracking_id
            )
            tracking.mark_as_opened(
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT')
            )
            return Response({'status': 'success'})
        except EmailTracking.DoesNotExist:
            return Response({'error': 'Invalid tracking ID'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'])
    def track_click(self, request, pk=None):
        """Отслеживание клика по ссылке"""
        tracking_id = request.GET.get('tracking_id')
        if not tracking_id:
            return Response({'error': 'Tracking ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            tracking = EmailTracking.objects.get(
                campaign_id=pk,
                tracking_id=tracking_id
            )
            tracking.mark_as_clicked(
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT')
            )
            return Response({'status': 'success'})
        except EmailTracking.DoesNotExist:
            return Response({'error': 'Invalid tracking ID'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'])
    def track_bounce(self, request, pk=None):
        """Отслеживание отказа письма"""
        tracking_id = request.GET.get('tracking_id')
        reason = request.data.get('reason', '')
        
        if not tracking_id:
            return Response({'error': 'Tracking ID is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            tracking = EmailTracking.objects.get(
                campaign_id=pk,
                tracking_id=tracking_id
            )
            tracking.mark_as_bounced(reason=reason)
            return Response({'status': 'success'})
        except EmailTracking.DoesNotExist:
            return Response({'error': 'Invalid tracking ID'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Получение статистики по кампаниям (корректные подсчёты по EmailTracking).

        Возвращаем агрегаты за последние 30 дней, чтобы согласовать с billing/dashboard.
        """
        from .models import EmailTracking
        from apps.mailer.models import Contact

        user = request.user
        now = timezone.now()
        last_30_days = now - timezone.timedelta(days=30)
        last_7_days = now - timezone.timedelta(days=7)
        prev_7_days = last_7_days - timezone.timedelta(days=7)

        # Базовый queryset трекинга по пользователю
        tracking_qs = EmailTracking.objects.filter(campaign__user=user)

        # Агрегаты за 30 дней
        sent_30 = tracking_qs.filter(sent_at__gte=last_30_days).count()
        opened_30 = tracking_qs.filter(sent_at__gte=last_30_days, opened_at__isnull=False).count()
        clicked_30 = tracking_qs.filter(sent_at__gte=last_30_days, clicked_at__isnull=False).count()

        # Тренды: сравниваем последние 7 дней и предыдущие 7 дней
        sent_7 = tracking_qs.filter(sent_at__gte=last_7_days).count()
        sent_prev_7 = tracking_qs.filter(sent_at__gte=prev_7_days, sent_at__lt=last_7_days).count()
        opened_7 = tracking_qs.filter(sent_at__gte=last_7_days, opened_at__isnull=False).count()
        opened_prev_7 = tracking_qs.filter(sent_at__gte=prev_7_days, sent_at__lt=last_7_days, opened_at__isnull=False).count()
        clicked_7 = tracking_qs.filter(sent_at__gte=last_7_days, clicked_at__isnull=False).count()
        clicked_prev_7 = tracking_qs.filter(sent_at__gte=prev_7_days, sent_at__lt=last_7_days, clicked_at__isnull=False).count()

        def calculate_trend(current, previous):
            if previous == 0:
                return 0
            return round(((current - previous) / previous) * 100, 1)

        # Новые подписчики (контакты) за 30 дней
        new_subscribers_30 = Contact.objects.filter(
            contact_list__owner=user,
            added_date__gte=last_30_days
        ).count()

        new_subscribers_7 = Contact.objects.filter(
            contact_list__owner=user,
            added_date__gte=last_7_days
        ).count()
        new_subscribers_prev_7 = Contact.objects.filter(
            contact_list__owner=user,
            added_date__gte=prev_7_days,
            added_date__lt=last_7_days
        ).count()

        return Response({
            'sent': sent_30,
            'opened': opened_30,
            'clicked': clicked_30,
            'sentTrend': calculate_trend(sent_7, sent_prev_7),
            'openedTrend': calculate_trend(opened_7, opened_prev_7),
            'clickedTrend': calculate_trend(clicked_7, clicked_prev_7),
            'newSubscribers': new_subscribers_30,
            'subscribersTrend': calculate_trend(new_subscribers_7, new_subscribers_prev_7),
        })

    @action(detail=True, methods=['post'])
    def retry(self, request, pk=None):
        """Повторная попытка отправки кампании"""
        campaign = self.get_object()

        # Проверяем, что кампания действительно в статусе ошибки
        if campaign.status != Campaign.STATUS_FAILED:
            return Response(
                {'detail': 'Повторная отправка возможна только для кампаний со статусом "Ошибка"'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Проверяем обязательные поля
        missing = []
        
        # Проверяем простые поля
        for field in ('name', 'template', 'sender_email', 'subject'):
            value = getattr(campaign, field)
            if not value:
                missing.append(field)
        
        # Проверяем sender_name через sender_email
        if not campaign.sender_email:
            missing.append('sender_email')
        
        # Проверяем списки контактов
        if not campaign.contact_lists.exists():
            missing.append('contact_lists')
            
        # Проверяем шаблон
        if not campaign.template:
            missing.append('template')
            
        # Проверяем email отправителя
        if not campaign.sender_email:
            missing.append('sender_email')

        if missing:
            return Response(
                {'detail': f'Не заполнены: {", ".join(missing)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Проверяем наличие контента в шаблоне
        if not campaign.template.html_content:
            return Response(
                {'detail': 'Шаблон не содержит HTML-контента'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Очищаем предыдущие записи об отправке
            EmailTracking.objects.filter(campaign=campaign).delete()
            CampaignRecipient.objects.filter(campaign=campaign).delete()

            # Обновляем статус и запускаем отправку
            campaign.status = Campaign.STATUS_SENDING
            campaign.save(update_fields=['status'])
            
            # Запускаем отправку через Celery
            from apps.campaigns.tasks import send_campaign
            task = send_campaign.delay(str(campaign.id))
            campaign.celery_task_id = task.id
            campaign.save(update_fields=['celery_task_id'])

            return Response({
                'detail': 'Кампания запущена повторно.',
                'status': 'sending',
                'status_display': campaign.get_status_display()
            })

        except Exception as e:
            campaign.status = Campaign.STATUS_FAILED
            campaign.save(update_fields=['status'])
            return Response(
                {'detail': f'Ошибка при повторной отправке: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def task_status(self, request, pk=None):
        """Проверка статуса задачи Celery"""
        campaign = self.get_object()
        
        if not campaign.celery_task_id:
            return Response({
                'task_id': None,
                'status': 'no_task',
                'message': 'Нет активной задачи'
            })
        
        try:
            from celery.result import AsyncResult
            task_result = AsyncResult(campaign.celery_task_id)
            
            # Получаем дополнительную информацию
            info = {
                'task_id': campaign.celery_task_id,
                'status': task_result.state,
                'ready': task_result.ready(),
                'successful': task_result.successful() if task_result.ready() else None,
                'failed': task_result.failed() if task_result.ready() else None,
            }
            
            # Если задача завершена, добавляем результат
            if task_result.ready():
                try:
                    result = task_result.result
                    info['result'] = result
                except Exception as e:
                    info['result_error'] = str(e)
            
            # Если задача в процессе, добавляем прогресс
            if task_result.state == 'PROGRESS':
                info['progress'] = task_result.info
            
            return Response(info)
            
        except Exception as e:
            return Response({
                'task_id': campaign.celery_task_id,
                'status': 'error',
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def system_status(self, request):
        """Проверка статуса системы Celery"""
        try:
            from celery import current_app
            inspect = current_app.control.inspect()
            
            # Проверяем активных workers
            stats = inspect.stats()
            active_tasks = inspect.active()
            registered_tasks = inspect.registered()
            
            # Проверяем очереди
            from django.core.cache import cache
            redis_info = {}
            try:
                import redis
                r = redis.Redis.from_url(settings.CELERY_BROKER_URL)
                redis_info = {
                    'connected': r.ping(),
                    'queue_lengths': {}
                }
                
                # Проверяем длину очередей
                for queue in ['campaigns', 'email', 'default']:
                    try:
                        redis_info['queue_lengths'][queue] = r.llen(queue)
                    except:
                        redis_info['queue_lengths'][queue] = 'unknown'
            except Exception as e:
                redis_info = {'error': str(e)}
            
            return Response({
                'workers': {
                    'count': len(stats) if stats else 0,
                    'active': bool(stats),
                    'details': stats
                },
                'tasks': {
                    'active': len(active_tasks.get('active', [])) if active_tasks else 0,
                    'registered': len(registered_tasks.get('registered', [])) if registered_tasks else 0
                },
                'redis': redis_info,
                'timestamp': timezone.now().isoformat()
            })
            
        except Exception as e:
            return Response({
                'error': 'Ошибка проверки статуса системы',
                'detail': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='export')
    def export_reports(self, request):
        """Экспорт отчетов по выбранным кампаниям в одном из форматов: CSV, XLSX, TXT, JSON.
        Body: { "campaign_ids": [..optional..], "format": "csv|xlsx|txt|json", "use_filtered": bool(optional) }
        Если campaign_ids пуст, можно экспортировать все кампании пользователя (или отфильтрованные на фронте и переданные списком).
        """
        data = request.data or {}
        campaign_ids = data.get('campaign_ids') or []
        export_format = (data.get('format') or 'csv').lower()
        statuses = data.get('statuses') or []  # optional list of statuses

        # Получаем queryset кампаний пользователя
        qs = self.get_queryset()
        if campaign_ids:
            qs = qs.filter(id__in=campaign_ids)
        elif statuses:
            qs = qs.filter(status__in=statuses)

        # Собираем агрегированные данные
        from apps.mailer.models import Contact as MailerContact
        items = []
        for c in qs:
            total_sent = EmailTracking.objects.filter(campaign=c).count()
            opens = EmailTracking.objects.filter(campaign=c, opened_at__isnull=False).count()
            clicks = EmailTracking.objects.filter(campaign=c, clicked_at__isnull=False).count()
            unsubscribed = MailerContact.objects.filter(
                id__in=EmailTracking.objects.filter(campaign=c).values_list('contact_id', flat=True),
                status=getattr(MailerContact, 'UNSUBSCRIBED', getattr(MailerContact, 'BLACKLIST', 'blacklist'))
            ).count()
            items.append({
                'id': str(c.id),
                'name': c.name,
                'subject': c.subject,
                'sender': f"{c.sender_name} <{c.sender_email.email if c.sender_email else ''}>",
                'created_at': c.created_at.isoformat(),
                'sent_at': c.sent_at.isoformat() if c.sent_at else None,
                'total_sent': total_sent,
                'opens': opens,
                'clicks': clicks,
                'unsubscribed': unsubscribed,
            })

        # Форматы
        if export_format == 'json':
            import json as _json
            response = HttpResponse(_json.dumps(items, ensure_ascii=False, indent=2), content_type='application/json; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="campaign_reports.json"'
            return response

        if export_format == 'txt':
            lines = []
            for it in items:
                lines.append(f"Кампания: {it['name'] or it['id']}\nТема: {it['subject'] or ''}\nОтправитель: {it['sender']}\nОтправлено: {it['total_sent']}\nОткрыли: {it['opens']}\nКликнули: {it['clicks']}\nОтписались: {it['unsubscribed']}\nДата отправки: {it['sent_at'] or ''}\n---")
            body = "\n\n".join(lines)
            response = HttpResponse(body, content_type='text/plain; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="campaign_reports.txt"'
            return response

        if export_format == 'xlsx':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                wb = Workbook()
                ws = wb.active
                ws.title = 'Reports'
                headers = ['ID', 'Кампания', 'Тема', 'Отправитель', 'Отправлено', 'Открыли', 'Кликнули', 'Отписались', 'Дата отправки']
                ws.append(headers)
                # Header style
                header_fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
                header_font = Font(bold=True, color='1E40AF')
                thin = Side(border_style='thin', color='E5E7EB')
                for col_idx, _ in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                for it in items:
                    ws.append([
                        it['id'], it['name'], it['subject'], it['sender'],
                        it['total_sent'], it['opens'], it['clicks'], it['unsubscribed'], it['sent_at'] or ''
                    ])
                # Zebra striping and borders
                stripe_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
                for r in range(2, ws.max_row + 1):
                    if r % 2 == 0:
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=r, column=c).fill = stripe_fill
                    for c in range(1, ws.max_column + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        if c in (5,6,7,8):
                            cell.alignment = Alignment(horizontal='center')
                # Auto column widths
                for column_cells in ws.columns:
                    length = max(len(str(column_cells[0].value or '')), *(len(str(cell.value or '')) for cell in column_cells))
                    adjusted = min(max(12, length + 2), 50)
                    ws.column_dimensions[column_cells[0].column_letter].width = adjusted
                # Freeze header
                ws.freeze_panes = 'A2'
                from io import BytesIO
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                response = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="campaign_reports.xlsx"'
                return response
            except Exception:
                # Фолбэк в CSV, если нет openpyxl
                export_format = 'csv'

        # CSV по умолчанию
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="campaign_reports.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Кампания', 'Тема', 'Отправитель', 'Отправлено', 'Открыли', 'Кликнули', 'Отписались', 'Дата отправки'])
        for it in items:
            writer.writerow([
                it['id'], it['name'], it['subject'], it['sender'],
                it['total_sent'], it['opens'], it['clicks'], it['unsubscribed'], it['sent_at'] or ''
            ])
        return response
    @action(detail=True, methods=['get'], url_path='export')
    def export_report(self, request, pk=None):
        """Экспорт отчета по кампании в CSV со счетчиками открытий, кликов и отписок."""
        campaign = self.get_object()
        # Собираем агрегаты
        total_sent = EmailTracking.objects.filter(campaign=campaign).count()
        opens = EmailTracking.objects.filter(campaign=campaign, opened_at__isnull=False).count()
        clicks = EmailTracking.objects.filter(campaign=campaign, clicked_at__isnull=False).count()
        from apps.mailer.models import Contact as MailerContact
        unsubscribed = MailerContact.objects.filter(
            id__in=EmailTracking.objects.filter(campaign=campaign).values_list('contact_id', flat=True),
            status=getattr(MailerContact, 'UNSUBSCRIBED', getattr(MailerContact, 'BLACKLIST', 'blacklist'))
        ).count()

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="campaign_report_{campaign.id}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Кампания', campaign.name or str(campaign.id)])
        writer.writerow(['Тема', campaign.subject or ''])
        writer.writerow(['Отправитель', f"{campaign.sender_name} <{campaign.sender_email.email if campaign.sender_email else ''}>"])
        writer.writerow([])
        writer.writerow(['Всего отправлено', total_sent])
        writer.writerow(['Открыли (кол-во)', opens])
        writer.writerow(['Кликнули (кол-во)', clicks])
        writer.writerow(['Отписались (кол-во)', unsubscribed])
        writer.writerow([])
        writer.writerow(['Email', 'Отправитель', 'Открыто', 'Кликов', 'Отписан'])

        # По каждому получателю
        trackings = EmailTracking.objects.filter(campaign=campaign).select_related('contact')
        unsubscribed_ids = set(MailerContact.objects.filter(
            id__in=trackings.values_list('contact_id', flat=True),
            status=getattr(MailerContact, 'UNSUBSCRIBED', getattr(MailerContact, 'BLACKLIST', 'blacklist'))
        ).values_list('id', flat=True))

        sender_mailbox = campaign.sender_email.email if campaign.sender_email else ''
        for t in trackings:
            writer.writerow([
                t.contact.email,
                sender_mailbox,
                1 if t.opened_at else 0,
                1 if t.clicked_at else 0,
                1 if t.contact_id in unsubscribed_ids else 0
            ])

        return response


class CampaignListView(LoginRequiredMixin, TemplateView):
    """
    Рендерит страницу «Список кампаний».
    SPA на Alpine.js сама подгрузит их через API.
    """
    template_name = 'campaigns_list.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_admin'] = self.request.user.is_staff
        
        # Проверяем наличие доменов и почт у пользователя
        from apps.emails.models import Domain, SenderEmail
        has_domains = Domain.objects.filter(owner=self.request.user).exists()
        has_emails = SenderEmail.objects.filter(owner=self.request.user).exists()
        context['has_domains_or_emails'] = has_domains or has_emails
        
        return context


class CampaignFormView(LoginRequiredMixin, TemplateView):
    """
    Рендерит страницу «Создать/редактировать кампанию».
    """
    template_name = 'campaigns_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_admin'] = self.request.user.is_staff
        
        # Проверяем наличие доменов и почт у пользователя
        from apps.emails.models import Domain, SenderEmail
        has_domains = Domain.objects.filter(owner=self.request.user).exists()
        has_emails = SenderEmail.objects.filter(owner=self.request.user).exists()
        context['has_domains_or_emails'] = has_domains or has_emails
        
        pk = kwargs.get('pk')
        
        if pk:
            try:
                campaign = Campaign.objects.get(pk=pk, user=self.request.user)
                # Преобразуем данные кампании в формат для фронтенда
                context['campaign'] = {
                    'id': str(campaign.id),
                    'name': campaign.name,
                    'subject': campaign.subject,
                    'sender_name': campaign.sender_email.sender_name if campaign.sender_email else '',
                    'reply_to': campaign.sender_email.reply_to if campaign.sender_email else '',
                    'status': campaign.status,
                    'scheduled_at': campaign.scheduled_at.isoformat() if campaign.scheduled_at else None,
                    'template': {
                        'id': str(campaign.template.id),
                        'title': campaign.template.title
                    } if campaign.template else None,
                    'sender_email': {
                        'id': str(campaign.sender_email.id),
                        'email': campaign.sender_email.email
                    } if campaign.sender_email else None,
                    'contact_lists': [
                        {
                            'id': str(cl.id),
                            'name': cl.name
                        } for cl in campaign.contact_lists.all()
                    ]
                }
            except Campaign.DoesNotExist:
                raise Http404("Кампания не найдена")
        
        return context


@require_GET
def track_email_open(request, campaign_id):
    """Обработчик открытия письма"""
    tracking_id = request.GET.get('tracking_id')
    if not tracking_id:
        raise Http404("Tracking ID is required")
        
    try:
        tracking = EmailTracking.objects.get(
            campaign_id=campaign_id,
            tracking_id=tracking_id
        )
        tracking.mark_as_opened(
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT')
        )
        # Return a 1x1 transparent pixel
        return HttpResponse(
            bytes.fromhex('47494638396101000100800000dbdbdb00000021f90401000000002c00000000010001000002024401003b'),
            content_type='image/gif'
        )
    except EmailTracking.DoesNotExist:
        raise Http404("Tracking record not found")

@require_GET
def track_email_click(request, campaign_id):
    """Обработчик клика по ссылке"""
    tracking_id = request.GET.get('tracking_id')
    url = request.GET.get('url')
    
    if not tracking_id:
        raise Http404("Tracking ID is required")
    if not url:
        raise Http404("URL parameter is required")
        
    try:
        tracking = EmailTracking.objects.get(
            campaign_id=campaign_id,
            tracking_id=tracking_id
        )
        tracking.mark_as_clicked(
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT')
        )
        
        # Redirect to the original URL
        return HttpResponseRedirect(url)
    except EmailTracking.DoesNotExist:
        raise Http404("Tracking record not found")


@require_GET
def unsubscribe(request, campaign_id):
    """Отписка контакта от рассылки: переводит контакт в черный список."""
    tracking_id = request.GET.get('tracking_id')
    if not tracking_id:
        raise Http404("Tracking ID is required")

    try:
        tracking = EmailTracking.objects.get(
            campaign_id=campaign_id,
            tracking_id=tracking_id
        )
        contact = tracking.contact
        # Помечаем контакт как черный список (или отписанный, если статус будет добавлен)
        from apps.mailer.models import Contact as MailerContact
        contact.status = getattr(MailerContact, 'UNSUBSCRIBED', getattr(MailerContact, 'BLACKLIST', 'blacklist'))
        contact.save(update_fields=['status'])
        # Возвращаем простую страницу подтверждения
        return HttpResponse("Вы успешно отписались от рассылки.")
    except EmailTracking.DoesNotExist:
        raise Http404("Tracking record not found")
